from django.shortcuts import render, redirect ,HttpResponse ,get_object_or_404
from .models import Shipment, Company
from .forms import ShipmentForm, CompanyForm ,ExcelUploadForm
#التالى للريبورت
import pandas as pd 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
#لرفع الاكسل
import os
from django.utils.dateparse import parse_date
from django.conf import settings
from datetime import datetime ,timedelta
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl import worksheet


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Shipment
import json
import io
import re

import pyexcel as pe
from django.shortcuts import render, redirect
from .models import Transaction
from .forms import TransactionForm
from django.db.models import Sum
from datetime import datetime
from django.http import HttpResponse
import openpyxl

@csrf_exempt
def save_shipment_changes(request):
    if request.method == "POST":
        data = json.loads(request.body)
        for entry in data['modifiedData']:
            shipment = Shipment.objects.get(id=entry['id'])
            setattr(shipment, entry['field'], entry['newValue'])
            shipment.save()
        return JsonResponse({"status": "success"})

# import django

# # تهيئة إعدادات Django (في حال لم تكن قد قمت بها في التطبيق)
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'your_project_name.settings')  # استبدل 'your_project_name' باسم مشروعك
# django.setup()

def convert_arabic_numbers_to_english(input_string):
    arabic_to_english = {'٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4', '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'}
    return ''.join(arabic_to_english.get(i, i) for i in input_string)

def convert_to_arabic_numbers(input_string):
    arabic_numbers = {'0': '٠', '1': '١', '2': '٢', '3': '٣', '4': '٤', '5': '٥', '6': '٦', '7': '٧', '8': '٨', '9': '٩'}
    return ''.join(arabic_numbers.get(i, i) for i in input_string)

def import_companies_and_shipments(file_path):

    # قراءة الملف كمجموعة من الشيتات
    excel_file = pd.ExcelFile(file_path)

    # استيراد الشركات والشحنات
    for sheet_name in excel_file.sheet_names:
        # الحصول على الرقم الضريبي (إذا كان موجودًا)
        tax_number = None # اتركه فارغًا إذا كنت لا تستخدمه
        company, created = Company.objects.get_or_create(name=sheet_name)

        # قراءة البيانات الخاصة بالشحنات من هذا الشيت
        df_shipments = pd.read_excel(file_path, sheet_name=sheet_name , header=1 )

        # طباعة أسماء الأعمدة للتأكد
        #print(f"أسماء الأعمدة في الشيت '{sheet_name}': {df_shipments.columns.tolist()}")
 
        for _, shipment_row in df_shipments.iterrows():
            # #إذا كان العمود 'اسم المركب' موجود
            # if 'اسم المركب' in shipment_row:
            #     vessel_name = shipment_row['اسم المركب']
            # else:
            #     vessel_name = None  # إذا لم يكن موجودًا، قم بتعيين قيمة افتراضية

            #إضافة الشحنات المرتبطة بالشركة
           

            Shipment.objects.create(
                    company = company,
                    ducmentsno = shipment_row['رقم الاقرار'],
                    agency = shipment_row ['اسم التوكيل'],
                    Acidno = shipment_row['ACID']if pd.notnull(shipment_row['ACID']) else 0,
                    NoCR = shipment_row['رقم الشهاده']if pd.notnull(shipment_row['رقم الشهاده']) else 0,
                    documents_received_date=datetime.strptime(shipment_row['ارسال الاوراق'], '%d.%m.%Y').date(),
                    expected_arrival_date=shipment_row['تاريخ الوصول'],
                    storge_data=datetime.strptime(shipment_row['تاريخ التخزين'],'%d.%m.%Y').date()if pd.notnull(shipment_row['تاريخ التخزين']) else '',
                    Delivery_data = datetime.strptime(shipment_row['سحب الاذن'],'%d.%m.%Y').date()if pd.notnull(shipment_row['سحب الاذن']) else '',
                    NoCE_data = datetime.strptime(shipment_row['فتح الشهاده'],'%d.%m.%Y').date()if pd.notnull(shipment_row['فتح الشهاده']) else '',
                    End_customs_data = datetime.strptime(shipment_row['الانتهاء'],'%d.%m.%Y').date()if pd.notnull(shipment_row['الانتهاء']) else '',
                    exchange_data = datetime.strptime(shipment_row['تاريخ الصرف'],'%d.%m.%Y').date()if pd.notnull(shipment_row['تاريخ الصرف']) else '',
                    vessel_name=shipment_row['اسم الباخره'] if pd.notnull(shipment_row['اسم الباخره']) else '',
                    bill_of_lading=shipment_row['رقم البوليصه'] if pd.notnull(shipment_row['رقم البوليصه']) else '',
                    weight=shipment_row['وزن الطرد'] if pd.notnull(shipment_row['وزن الطرد']) else '',
                    packages_count=shipment_row['عدد الطرود'] if pd.notnull(shipment_row['عدد الطرود']) else '',
                    invoice_number=shipment_row['رقم الفاتوره'] if pd.notnull(shipment_row['رقم الفاتوره']) else '',
                    # دمج الملاحظات مع تحويل الأرقام
                    comment = (
                        (shipment_row['ملاحظات 1'] if pd.notnull(shipment_row['ملاحظات 1']) else '') + '\n' +
                        (shipment_row['ملاحظات'] if pd.notnull(shipment_row['ملاحظات']) else '') + '\n' +
                        (shipment_row['ملاحظات 12'] if pd.notnull(shipment_row['ملاحظات 12']) else '') + '\n' +
                        (shipment_row['ملاحظات 2'] if pd.notnull(shipment_row['ملاحظات 2']) else '')
                    )
            )

    print("تم استيراد البيانات بنجاح!")

def upload_excel(request):
    
    if request.method == 'POST' and request.FILES['excel_file']:
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # مسار حفظ الملف
            file_path = os.path.join(settings.MEDIA_ROOT, 'uploaded_files', excel_file.name)
            
            # التأكد من وجود المجلد وإنشاؤه إذا كان مفقودًا
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, 'wb+') as f:
                for chunk in excel_file.chunks():
                    f.write(chunk)
            
            # استيراد البيانات من الملف
            import_companies_and_shipments(file_path)
            
            return redirect('home')  # يمكن استبدال هذه بـ URL عرض رسالة النجاح أو إعادة التوجيه إلى صفحة أخرى.
    else:
        form = ExcelUploadForm()

    return render(request, 'upload_excel.html', {'form': form})

def home(request):
    companies = Company.objects.all()
    shipments = Shipment.objects.all()



    return render(request,'index.html',{'companies': companies ,'shipments': shipments})

def export_shipments_to_excel(request):
    # منطق تصدير البيانات
    shipments = Shipment.objects.all()
    grouped_shipments = {}

    # تجميع الشحنات حسب الشركة
    for shipment in shipments:
        company_name = shipment.company.name  # على افتراض وجود حقل name في نموذج الشركة
        if company_name not in grouped_shipments:
            grouped_shipments[company_name] = []
        grouped_shipments[company_name].append(shipment)

    # إعداد استجابة لتحميل ملف Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=shipments.xlsx'  # تعيين اسم الملف عند التحميل

    # كتابة البيانات إلى ملف Excel باستخدام pandas
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        for company_name, shipments in grouped_shipments.items():
            data = []
            for shipment in shipments:
                data.append({
                    "رقم الاقرار": shipment.ducmentsno,
                    "التوكيل ": shipment.agency,
                    "ACID": shipment.Acidno,
                    "رقم الشهاده": shipment.NoCR,
                    "استلام الاوراق": shipment.documents_received_date,
                    "تاريخ الوصول": shipment.expected_arrival_date,
                    "تاريخ التخزين": shipment.storge_data,
                    "سحب الاذن": shipment.Delivery_data,
                    "فتح الشهاده": shipment.NoCE_data,
                    "الانتهاء": shipment.End_customs_data,
                    "الصرف": shipment.exchange_data,
                    "اسم الباخره": shipment.vessel_name,
                    "رقم البوليصه": shipment.bill_of_lading,
                    "الوزن": shipment.weight,
                    "عدد الطرود": shipment.packages_count,
                    "رقم الفاتره": shipment.invoice_number,
                    "ملاحظات": shipment.comment
                })
            
            # تحويل البيانات إلى DataFrame ثم تصديرها إلى Excel
            df = pd.DataFrame(data)
            
            # تحويل بيانات إلى Excel وكتابة التنسيق
            df.to_excel(writer, sheet_name=company_name, index=False)

            # الحصول على الورقة التي تم إنشاؤها
            worksheet = writer.sheets[company_name]

            # تعيين الاتجاه من اليمين لليسار
            worksheet.sheet_view.rigtToleft = True




            # إعداد تنسيق الخلايا
            for row in worksheet.iter_rows():
                for cell in row:
                    # محاذاة النص من اليمين لليسار
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            # ضبط عرض الأعمدة تلقائيًا بناءً على المحتوى
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # الحصول على اسم العمود
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width


            # إضافة جدول داخل الورقة
            # استبدال المسافات في اسم الجدول بشرطة سفلية (underscore)
            table_name = f"Table_{company_name.replace(' ', '_')}"
            table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
            table = Table(displayName=table_name, ref=table_ref)
            # تخصيص نمط الجدول
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=True
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
            

    return response

def add_company(request):
    if request.method == "POST":
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('company_list')
    else:
        form = CompanyForm()
    return render(request, 'add_company.html', {'form': form})

# عرض الشركات
def company_list(request):
    companies = Company.objects.all()
    return render(request, 'company_list.html', {'companies': companies})

# عرض نموذج إضافة شحنة
def add_shipment(request):
    if request.method == "POST":
        form = ShipmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('shipment_list')
    else:
        form = ShipmentForm()
    return render(request, 'add_shipment.html', {'form': form})

# عرض الشحنات
def shipment_list(request):
    
    companies = Company.objects.all()
    shipments = Shipment.objects.all()
    return render(request, 'shipment_list.html', {'companies': companies ,'shipments': shipments.order_by('company')})


def convert_to_minutes(value):
    if pd.isna(value):
        return 0
    try:
        match = re.match(r'^(\d{1,2}):(\d{2})$', str(value).strip())
        if match:
            hours, minutes = map(int, match.groups())
            return hours * 60 + minutes
        else:
            return int(value)
    except:
        return 0

def format_minutes_arabic(total_minutes):
    #days = total_minutes // (8 * 60)
    hours = total_minutes  // 60
    minutes = total_minutes % 60
    return f" {hours} ساعه {minutes} دقيقه" 

def upload_and_generate_report(request):
    if request.method == 'POST' and request.FILES.get('file'):
        # استلام التواريخ المدخلة من المستخدم
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        # تحويل التواريخ إلى كائنات datetime
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            return HttpResponse("تاريخ البداية أو النهاية غير صحيح، يرجى إدخال التواريخ بشكل صحيح.", status=400)

        file = request.FILES['file']
        filename = file.name.lower()

        # 👇 قراءة حسب نوع الملف
        if filename.endswith('.xls'):
            sheet = pe.get_sheet(file_type='xls', file_content=file.read())
            df = pd.DataFrame(sheet.to_array()[1:], columns=sheet.row[0])
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(file, engine='openpyxl')
        else:
            return HttpResponse("⚠️ نوع الملف غير مدعوم. الرجاء رفع ملف .xls أو .xlsx")

        df['Date'] = pd.to_datetime(df['Date'], dayfirst=True)

        all_days = pd.date_range(start=start_date, end=end_date, freq='D')
        working_days = [d for d in all_days if d.weekday() != 4]

        employee_names = df['Name'].dropna().unique()
        report_data = []

        for date in working_days:
            for name in employee_names:
                record = df[(df['Name'] == name) & (df['Date'] == date)]
                if not record.empty:
                    clock_in = record.iloc[0].get('Clock In', '')
                    clock_out = record.iloc[0].get('Clock Out', '')
                    Late = record.iloc[0].get('Late', '')
                    Early = record.iloc[0].get('Early', '') 
                    OT_Time = record.iloc[0].get('OT Time', '')

                    # إذا كانت بيانات "الدخول" و "الخروج" مفقودة، يعتبر الموظف غائبًا
                    if not clock_in and not clock_out:
                        status = 'غائب'
                    elif Early != '' and Late != '':
                        status = 'تاخير/انصراف'
                    elif Late is not None and Late != '':
                        status = 'تأخير'
                    elif Early is not None and Early != '':
                        status = 'انصراف مبكر'
                    elif not clock_in and clock_out:
                        status = 'بدون دخول'
                    elif clock_in and not clock_out:
                        status = 'بدون خروج'
                    else:
                        status = 'حضور'
                else:
                    clock_in = ''
                    clock_out = ''
                    Late = ''
                    Early = ''
                    OT_Time = ''
                    status = 'غائب'

                report_data.append({
                    'اليوم': date.strftime('%A'),
                    'التاريخ': date.strftime('%Y-%m-%d'),
                    'الموظف': name,
                    'الدخول': clock_in,
                    'الخروج': clock_out,
                    'التاخير': Late,
                    'الانصراف المبكر' : Early,
                    'الوقت الاضافى' : OT_Time ,

                    'الحالة': status
                })

        report_df = pd.DataFrame(report_data)
        report_df = report_df.sort_values(by=['الموظف', 'التاريخ'])

        # الملخص
        summary_data = []
        for name, group in report_df.groupby('الموظف'):
            total_minutes = group['التاخير'].apply(convert_to_minutes).sum()
            total_minutes_early = group['الانصراف المبكر'].apply(convert_to_minutes).sum()
            total_minutes_OT = group['الوقت الاضافى'].apply(convert_to_minutes).sum()

            total_Late = (group['التاخير'] != '').sum()
            total_early = (group['الانصراف المبكر'] != '').sum()
            total_absent = (group['الحالة'] == 'غائب').sum()

            summary_data.append({
                'الموظف': name,
                'إجمالي التأخير': format_minutes_arabic(total_minutes),
                'إجمالي الانصراف المبكر': format_minutes_arabic(total_minutes_early),
                'الوقت الاضافى': format_minutes_arabic(total_minutes_OT),
                'عدد أيام الغياب': total_absent,
                'عدد أيام التاخير': total_Late,
                'عدد أيام الانصراف': total_early
            })

        summary_df = pd.DataFrame(summary_data)

        #حفظ التقرير داخل ملف في الذاكرة
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            report_df.to_excel(writer, sheet_name='بيانات الحضور', index=False)
            summary_df.to_excel(writer, sheet_name='الملخص', index=False)

        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        start_str = start_date.strftime('%Y-%m-%d')
        end_str = end_date.strftime('%Y-%m-%d')

        response['Content-Disposition'] = F'attachment; filename=finger-report_from{start_str}to{end_str}.xlsx'
        return response
        
        
        # buffer = io.BytesIO()
        # with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        #     report_df.to_excel(writer, sheet_name='بيانات الحضور', index=False)
        #     summary_df.to_excel(writer, sheet_name='الملخص', index=False)

        #     # تعديل حجم الأعمدة تلقائيًا
        #     workbook = writer.book
        #     for sheet_name in writer.sheets:
        #         worksheet = writer.sheets[sheet_name]
        #         for column_cells in worksheet.columns:
        #             max_length = 0
        #             column_letter = column_cells[0].column_letter
        #             for cell in column_cells:
        #                 if cell.value:
        #                     max_length = max(max_length, len(str(cell.value)))
        #             worksheet.column_dimensions[column_letter].width = max_length + 0

        # buffer.seek(0)
        # response = HttpResponse(
        #     buffer.read(),
        #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )

        # start_str = start_date.strftime('%Y-%m-%d')
        # end_str = end_date.strftime('%Y-%m-%d')

        # response['Content-Disposition'] = f'attachment; filename=finger-report_from{start_str}_to_{end_str}.xlsx'
        # return response

    return render(request, 'upload.html')

from .models import Transaction
from .forms import TransactionForm
from django.db.models import Sum
from datetime import datetime
from django.http import HttpResponse
import openpyxl

def transaction_list(request):
    form = TransactionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('transaction_list')

    transactions = Transaction.objects.all()

    # فلترة التاريخ (لو محدد)
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')
    if from_date and to_date:
        transactions = transactions.filter(date__range=[from_date, to_date])

    currencies = ['EGP', 'USD', 'EUR']
    data_by_currency = {}
    for curr in currencies:
        txs = transactions.filter(currency=curr)
        total_in = txs.filter(type='in').aggregate(Sum('amount'))['amount__sum'] or 0
        total_out = txs.filter(type='out').aggregate(Sum('amount'))['amount__sum'] or 0
        balance = total_in - total_out
        data_by_currency[curr] = {
            'transactions': txs,
            'in': total_in,
            'out': total_out,
            'balance': balance
        }

    return render(request, 'transactions.html', {
        'form': form,
        'data_by_currency': data_by_currency,
        'from_date': from_date,
        'to_date': to_date
    })

def export_transactions_excel(request):
    transactions = Transaction.objects.all()
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')
    if from_date and to_date:
        transactions = transactions.filter(date__range=[from_date, to_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الحركات"
    headers = ['التاريخ', 'البيان', 'النوع', 'العملة', 'المبلغ']
    ws.append(headers)

    for tx in transactions:
        ws.append([str(tx.date), tx.description, tx.get_type_display(), tx.currency, float(tx.amount)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
    wb.save(response)
    return response
